# analytics_slot_overlay.py - FIXED VERSION
"""ANALYTICS SLOT OVERLAY - Wiring Analytics Brain to Live Execution (STABLE DATE DETECTION)"""
import json
import pandas as pd
from pathlib import Path
from datetime import datetime
import sys
from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore')

class AnalyticsSlotOverlay:
    def __init__(self):
        self.base_dir = Path(__file__).resolve().parent
        self.latest_date = None
        self.slots = ["FRBD", "GZBD", "GALI", "DSWR"]
        self.final_stakes = {}
        self.final_total = 0
        
    def discover_latest_date(self):
        """Discover latest date from bet_plan_master_*.xlsx (FIXED VERSION)"""
        bet_engine_dir = self.base_dir / "predictions" / "bet_engine"
        if not bet_engine_dir.exists():
            print("❌ Bet engine directory not found")
            return False
        
        # 🔥 CRITICAL FIX: Use bet_plan_master_*.xlsx instead of final_bet_plan_*.xlsx
        master_files = list(bet_engine_dir.glob("bet_plan_master_*.xlsx"))
        
        if not master_files:
            print("❌ No bet_plan_master_*.xlsx files found in bet_engine directory")
            return False
            
        # Extract dates only from bet_plan_master files
        master_dates = set()
        for file in master_files:
            date_str = file.stem.split('_')[-1]
            if date_str.isdigit() and len(date_str) == 8:
                master_dates.add(date_str)
        
        if not master_dates:
            print("❌ No valid dates found in bet_plan_master files")
            return False
            
        self.latest_date = max(master_dates)
        print(f"📅 Detected latest date (from bet_plan_master): {self.latest_date}")
        return True

    def ensure_final_bet_plan_exists(self):
        """
        Ensure final_bet_plan_{latest_date}.xlsx exists.
        If missing, synthesize it from bet_plan_master_{latest_date}.xlsx
        with a minimal 'final_slot_plan' sheet.
        """
        bet_engine_dir = self.base_dir / "predictions" / "bet_engine"
        final_file = bet_engine_dir / f"final_bet_plan_{self.latest_date}.xlsx"
        master_file = bet_engine_dir / f"bet_plan_master_{self.latest_date}.xlsx"

        if final_file.exists():
            return final_file

        if not master_file.exists():
            print("❌ bet_plan_master file not found, cannot synthesize final_bet_plan")
            return None

        try:
            # Load master summary to get per-slot total stake
            df_summary = pd.read_excel(master_file, sheet_name="summary")
            df_summary.columns = [str(c).strip().lower() for c in df_summary.columns]

            slot_col = None
            total_stake_col = None
            for col in df_summary.columns:
                if col == "slot":
                    slot_col = col
                if col in ("total_stake", "total_slot_stake", "slot_total_stake"):
                    total_stake_col = col

            if not slot_col or not total_stake_col:
                print("❌ Could not find slot/total_stake columns in master summary")
                return None

            rows = []
            for slot in self.slots:
                slot_rows = df_summary[df_summary[slot_col] == slot]
                if not slot_rows.empty:
                    stake_val = float(slot_rows[total_stake_col].iloc[0])
                else:
                    stake_val = 0.0
                rows.append({"slot": slot, "final_slot_stake": stake_val})

            df_final = pd.DataFrame(rows)

            # Save minimal final_bet_plan with single 'final_slot_plan' sheet
            final_file.parent.mkdir(parents=True, exist_ok=True)
            with pd.ExcelWriter(final_file, engine="openpyxl") as writer:
                df_final.to_excel(writer, sheet_name="final_slot_plan", index=False)

            print(f"✅ Synthesized final_bet_plan from master: {final_file}")
            return final_file
        except Exception as e:
            print(f"❌ Error synthesizing final_bet_plan: {e}")
            return None

    def load_analytics_data(self):
        """Load analytics JSON files with EXACT structures"""
        analytics_data = {
            'dynamic_stakes': {},
            'money_management': {},
            'smart_weights': {}
        }
        
        # 1) dynamic_stake_plan.json - EXACT STRUCTURE
        dynamic_file = self.base_dir / "logs" / "performance" / "dynamic_stake_plan.json"
        if dynamic_file.exists():
            try:
                with open(dynamic_file, 'r') as f:
                    data = json.load(f)

                stakes_obj = (
                    data.get("stakes")
                    or data.get("slot_stakes")
                    or data.get("base_slot_stakes")
                    or {}
                )
                analytics_data['dynamic_stakes'] = {
                    slot: float(stakes_obj.get(slot, 0) or 0) for slot in self.slots
                }
                derived_total = sum(analytics_data['dynamic_stakes'].values())
                analytics_data['dynamic_total'] = float(
                    data.get("total_daily_stake", derived_total) or derived_total
                )
                print(f"✅ Loaded dynamic stakes (from JSON): {analytics_data['dynamic_stakes']}")
            except Exception as e:
                print(f"⚠️ Error loading dynamic_stake_plan: {e}")
                return None
        else:
            print("⚠️ dynamic_stake_plan.json not found")
            return None
            
        # 2) money_management_plan.json - EXACT STRUCTURE  
        money_file = self.base_dir / "logs" / "performance" / "money_management_plan.json"
        if money_file.exists():
            try:
                with open(money_file, 'r') as f:
                    data = json.load(f)
                
                # EXACT JSON STRUCTURE: data["daily_limits"] and data["bankroll_rules"]
                daily_limits = data.get("daily_limits", {})
                bankroll_rules = data.get("bankroll_rules", {})
                
                max_total = float(daily_limits.get("max_total_stake", 300))
                max_single = float(daily_limits.get("max_single_stake", 150))
                recommended_risk = float(bankroll_rules.get("recommended_daily_risk", 500))
                
                analytics_data['money_management'] = {
                    'max_total_stake': max_total,
                    'max_single_stake': max_single,
                    'recommended_daily_risk': recommended_risk,
                }
                print(f"✅ Loaded money management: max_total={max_total}, max_single={max_single}, recommended_daily_risk={recommended_risk}")
            except Exception as e:
                print(f"⚠️ Error loading money_management_plan: {e}")
                # Use defaults but continue
                analytics_data['money_management'] = {
                    'max_total_stake': 300,
                    'max_single_stake': 150,
                    'recommended_daily_risk': 500,
                }
        else:
            print("⚠️ money_management_plan.json not found — using defaults")
            analytics_data['money_management'] = {
                'max_total_stake': 300,
                'max_single_stake': 150,
                'recommended_daily_risk': 500,
            }
            
        # 3) smart_fusion_weights.json - EXACT STRUCTURE
        weights_file = self.base_dir / "logs" / "performance" / "smart_fusion_weights.json"
        if weights_file.exists():
            try:
                with open(weights_file, 'r') as f:
                    data = json.load(f)
                
                # EXACT JSON STRUCTURE: direct slot keys at root level
                if isinstance(data, dict):
                    slot_weights = data
                else:
                    slot_weights = {}
                    
                analytics_data['smart_weights'] = {
                    slot: float(slot_weights.get(slot, 1.0)) for slot in self.slots
                }
                print(f"✅ Loaded smart weights: {analytics_data['smart_weights']}")
            except Exception as e:
                print(f"⚠️ Error loading smart_fusion_weights: {e}")
                analytics_data['smart_weights'] = {slot: 1.0 for slot in self.slots}
        else:
            analytics_data['smart_weights'] = {slot: 1.0 for slot in self.slots}
                
        return analytics_data

    def calculate_safe_stakes(self, analytics_data):
        """Calculate safe stakes with money management limits"""
        raw_stakes = analytics_data['dynamic_stakes']
        money_mgmt = analytics_data['money_management']
        
        max_total = money_mgmt.get('max_total_stake', 300)
        max_single = money_mgmt.get('max_single_stake', 150)
        recommended_risk = money_mgmt.get('recommended_daily_risk', 500)
        
        raw_total = sum(raw_stakes.values())
        
        print(f"\n[Analytics Overlay] RAW slot stakes: {raw_stakes} (total={raw_total})")
        print(f"[Analytics Overlay] LIMITS: max_total={max_total}, max_single={max_single}, recommended_daily_risk={recommended_risk}")
        
        if raw_total <= 0:
            print("❌ No positive dynamic stakes, skipping overlay")
            return None
            
        # Calculate safe total stake
        safe_total = min(raw_total, max_total, recommended_risk)
        scale_factor = safe_total / raw_total if raw_total > 0 else 1.0
        
        # Calculate final stakes per slot
        final_stakes = {}
        for slot in self.slots:
            scaled = raw_stakes[slot] * scale_factor
            scaled_rounded = round(scaled)
            final_stakes[slot] = min(scaled_rounded, max_single)
            
        final_total = sum(final_stakes.values())
        
        print(f"[Analytics Overlay] FINAL slot stakes: {final_stakes} (total={final_total})")
        return final_stakes, final_total

    def _find_header_row_and_columns(self, sheet):
        """Smart header row detection with robust slot/stake mapping (fixed)."""
        STAKE_KEYWORDS = [
            'final_slot_stake',
            'slot_total_stake',
            'slot_stake',
            'slot_total',
            'stake',
        ]

        header_row_idx = None
        slot_col = None
        stake_col = None

        max_rows_to_scan = min(10, sheet.max_row)

        for row_idx in range(1, max_rows_to_scan + 1):
            current_slot_col = None
            current_stake_col = None

            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if not cell_value:
                    continue

                cell_str = str(cell_value).strip().lower()
                norm = cell_str.replace(" ", "_")

                # SLOT COLUMN DETECTION (EXACT)
                # We only treat these as slot headers:
                #   - "slot"
                #   - "slot_name"
                #   - "slot_id"
                # This avoids picking "slot_confidence_level" etc.
                if norm in ("slot", "slot_name", "slot_id"):
                    current_slot_col = col_idx

                # STAKE COLUMN DETECTION (PREFER FINAL SLOT STAKE)
                for keyword in STAKE_KEYWORDS:
                    if keyword in norm:
                        current_stake_col = col_idx
                        break

            # If both found in this row, we found the header row
            if current_slot_col and current_stake_col:
                header_row_idx = row_idx
                slot_col = current_slot_col
                stake_col = current_stake_col
                print(f"✅ Found header row {row_idx}: slot_col={slot_col}, stake_col={stake_col}")
                break

        return header_row_idx, slot_col, stake_col

    def update_final_bet_plan(self, final_stakes):
        """Update final_bet_plan Excel file with smart header detection"""

        final_file = self.ensure_final_bet_plan_exists()
        if final_file is None or not final_file.exists():
            print("❌ Final bet plan file not available, cannot apply overlay")
            return False

        try:
            # Use openpyxl to preserve formatting
            workbook = load_workbook(final_file)
            sheet_name = None
            for name in workbook.sheetnames:
                if 'final_slot_plan' in name.lower():
                    sheet_name = name
                    break
                    
            if not sheet_name:
                print("❌ final_slot_plan sheet not found")
                return False
                
            sheet = workbook[sheet_name]
            
            # Smart header detection
            header_row_idx, slot_col, stake_col = self._find_header_row_and_columns(sheet)
            
            if not header_row_idx or not slot_col or not stake_col:
                print("❌ Could not detect header row with slot + stake columns")
                return False
                
            # Update slot stakes (data starts from header_row_idx + 1)
            updated_count = 0
            for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
                slot_cell = sheet.cell(row=row_idx, column=slot_col)
                slot_value = slot_cell.value
                
                if slot_value in self.slots:
                    old_stake = sheet.cell(row=row_idx, column=stake_col).value
                    new_stake = final_stakes.get(slot_value, old_stake)
                    sheet.cell(row=row_idx, column=stake_col).value = new_stake
                    print(f"   Updated {slot_value}: {old_stake} → {new_stake}")
                    updated_count += 1
                    
            # Save the workbook
            workbook.save(final_file)
            print(f"✅ Updated {updated_count} slots in final_bet_plan")
            return True
            
        except Exception as e:
            print(f"❌ Error updating final_bet_plan: {e}")
            return False

    def update_execution_readiness(self, final_total):
        """Update execution_readiness_summary.json"""
        exec_file = self.base_dir / "logs" / "performance" / "execution_readiness_summary.json"
        if not exec_file.exists():
            print("⚠️ execution_readiness_summary.json not found, skipping update")
            return True
            
        try:
            with open(exec_file, 'r') as f:
                data = json.load(f)
                
            # Update base totals
            old_total = data.get('base_final_total_stake', 0)
            data['base_final_total_stake'] = float(final_total)
            
            # Update recommended real stake based on mode
            mode = data.get('mode', 'PAPER_ONLY')
            stake_multiplier = data.get('stake_multiplier', 0.0)
            if mode in ['GO_LIVE_FULL', 'GO_LIVE_LIGHT']:
                data['recommended_real_total_stake'] = final_total * stake_multiplier
                
            # Add note
            notes = data.get('notes', [])
            stake_str = "/".join([str(int(v)) for v in self.final_stakes.values()])
            notes.append(f"Analytics Overlay: final_slot_stakes {stake_str} (total={final_total})")
            data['notes'] = notes[-5:]  # Keep last 5 notes
            
            with open(exec_file, 'w') as f:
                json.dump(data, f, indent=2)
                
            print(f"✅ Updated execution_readiness: base_final_total={old_total} → {final_total}")
            return True
            
        except Exception as e:
            print(f"⚠️ Error updating execution_readiness: {e}")
            return True  # Don't fail the whole process

    def run_overlay(self):
        """Run complete analytics overlay"""
        print("🧮 ANALYTICS SLOT OVERLAY - Wiring Analytics to Live Execution")
        print("=" * 60)
        
        # Step 1: Discover latest date
        if not self.discover_latest_date():
            print("⚠️ No bet plan detected – skipping overlay.")
            return True
            
        # Step 2: Load analytics data
        analytics_data = self.load_analytics_data()
        if not analytics_data or not analytics_data.get('dynamic_stakes'):
            print("⚠️ No analytics data available, skipping overlay")
            return True
            
        # Step 3: Calculate safe stakes
        result = self.calculate_safe_stakes(analytics_data)
        if not result:
            print("⚠️ Unable to compute safe stakes – skipping overlay.")
            return True
            
        self.final_stakes, self.final_total = result
        
        # Step 4: Update final bet plan
        if not self.update_final_bet_plan(self.final_stakes):
            print("⚠️ Unable to update final bet plan – overlay completed with warnings.")
            return True
            
        # Step 5: Update execution readiness
        self.update_execution_readiness(self.final_total)
        
        print("✅ ANALYTICS OVERLAY COMPLETED SUCCESSFULLY")
        return True

def main():
    overlay = AnalyticsSlotOverlay()
    success = overlay.run_overlay()
    return 0 if success else 1

if __name__ == "__main__":
    exit(main())